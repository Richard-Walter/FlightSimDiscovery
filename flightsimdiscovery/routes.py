import os, sys
import secrets
from openpyxl import Workbook, load_workbook
import json
from PIL import Image
from flask import render_template, url_for, flash, redirect, request, abort
from flightsimdiscovery import app, db, bcrypt
from flightsimdiscovery.forms import RegistrationForm, LoginForm, UpdateAccountForm, PoiCreateForm, PoiUpdateForm
from flightsimdiscovery.models import User, Pois, Ratings, Favorites, Visited
from flask_login import login_user, current_user, logout_user, login_required
from utilities import get_country_region, get_country_list, get_region_list, get_category_list, region_details, countries_details

# example data that needs to be created from database and then posted to home.html
# data = [
#   { 'name': 'Home', 'category': 'Bush airport', 'country': 'Australia', 'description': 'Awesome remote bush strip', 'rating': '5','icon': 'http://maps.google.com/mapfiles/ms/micons/pink-pushpin.png', 'lat': -34.44315867450577, 'lng': 150.84022521972656 },
#   { 'name': 'Work', 'category': 'Sea base', 'country': 'Vietnam', 'description': 'Great water landing', 'rating': '4', 'icon': 'http://maps.google.com/mapfiles/ms/micons/pink-pushpin.png', 'lat': -35.284, 'lng': 150.833 },
#   { 'name': 'Airport', 'category': 'National Park', 'country': 'Congo', 'description': 'Giant trees in mountains areas', 'rating': '3', 'icon': '/static/img/marker/map-mark.png', 'lat': -35.123, 'lng': 150.534 },
# ]

@app.route("/", methods=['GET', 'POST'])
@app.route("/index", methods=['GET', 'POST'])
@app.route("/home", methods=['GET', 'POST'])
def home():


    #variables required for google maps to display data
    map_data = []
    map_data_dict = {}
    map_init = {'zoom': 3, 'lat':23.6, 'long':170.9} # centre of map
    pois = Pois.query.all()
    # user_pois_list = []
    user_pois_dict = {}
    search_defaults = {'Category': 'Category', 'Region': 'Region', 'Country': 'Country', 'Rating':'Rating'} 
    is_authenticated = False
    user_id = None

    if current_user.is_authenticated:

        is_authenticated = True

        # Create a list of Users POIS for the google map info window to use
        user_id = current_user.id
        user_pois = Pois.query.filter_by(user_id=user_id).all() #  returns a list

        for poi in user_pois:

            rating = str(get_user_rating(poi.id))  
            visited= get_visited(poi.id)
            favorited = get_favorited(poi.id)

            user_pois_dict[poi.id] = {'user_rating': rating, 'visited': visited,'favorited': favorited}        

    # check if user has submitted a search and filter database
    if request.method == 'POST':

        if 'search_form_submit' in request.form:
        
            category = request.form.get('selectCategory').strip()
            region = request.form.get('selectRegion').strip()
            country = request.form.get('selectCountry').strip()
            rating = request.form.get('selectRating').strip()

            #update search defaults so it shows the last search criteria
            search_defaults['Category']=category
            search_defaults['Region']=region
            search_defaults['Country']=country
            search_defaults['Rating']=rating

            # Creat the map intit variables
            if country != 'Country':

                map_init['zoom'] = 6    # default country zoom
                map_init['lat'] = countries_details[country][1]
                map_init['long'] = countries_details[country][2]

            elif region != 'Region':
                map_init['zoom'] = region_details[region][2]
                map_init['lat'] = region_details[region][0]
                map_init['long'] = region_details[region][1]

            # pois = Pois.query.filter_by(region='Oceania')
            if category != 'Category':
                pois = Pois.query.filter_by(category=category)
            if region != 'Region':
                pois = Pois.query.filter_by(region=region)
            if country != 'Country':
                pois = Pois.query.filter_by(country=country)         
            if rating != 'Rating':
                pois = Pois.query.filter_by(rating=rating)

        elif 'ratingOptions' in request.form:
            
            #  Stores users POI preferences from submitted form
            rating_score = request.form.get('ratingOptions') 
            poi_id = request.form.get('poi_id')
                           
            # Update ratings table
            if rating_score:
                rating = Ratings.query.filter_by(user_id=user_id).filter_by(poi_id=poi_id).first()
                print('\n\n')
                print('OLD RATING: ', rating)
                if rating:  # update rating score

                    rating.rating_score=rating_score
                else:
                    rating = Ratings(user_id=user_id, poi_id= poi.id, rating_score=rating_score)
                    db.session.add(rating)
                db.session.commit()
                print('NEW RATING: ', rating)

        elif 'favoriteChecked' in request.form:
            
            #  Stores users POI preferences from submitted form
            favorited = request.form.get('favoriteChecked')
            poi_id = request.form.get('poi_id')

            # Add/Update favorites table
            if favorited:
                favorite = Favorites.query.filter_by(user_id=user_id).filter_by(poi_id=poi_id).first()
                print('OLD favorite: ', favorite)

                if favorite: # record already exists do nothing
                    pass
                else:
                    favorite=Favorites(user_id=user_id, poi_id= poi_id)
                    db.session.add(favorite)
                    db.session.commit()
                    print('NEW favorite: ', favorite)
            else:   # remove record from db if exists
                favorite = Favorites.query.filter_by(user_id=user_id).filter_by(poi_id=poi_id).first()

                if favorite: 
                    print('REMOVING favorite: ', favorite)
                    db.session.delete(favorite)
                    db.session.commit()
                               
        elif 'visitedChecked' in request.form:
            
            #  Stores users POI preferences from submitted form
            visited = request.form.get('visitedChecked')
            poi_id = request.form.get('poi_id')
                               
            # Add/Update Visited table
            if visited:
                visit = Visited.query.filter_by(user_id=user_id).filter_by(poi_id=poi_id).first()
                print('OLD Visited: ', visit)

                if visit: # record already exists do nothing
                    pass
                else:
                    visit=Visited(user_id=user_id, poi_id= poi_id)
                    db.session.add(visit)
                    db.session.commit()
                    print('NEW Visited: ', visit)
            else:   # remove record from db if exists
                visit = Visited.query.filter_by(user_id=user_id).filter_by(poi_id=poi_id).first()

                if visit: 
                    print('REMOVING Visited: ', visit)
                    db.session.delete(visit)
                    db.session.commit()

            # return   # dont wont to reload the page, just store the users settg


    #create the Point of Interest dictionary that gets posted for map to use
    for poi in pois:
        # print('Poi', poi)
        data_dic = {}
        data_dic['id'] = poi.id
        data_dic['user_id'] = poi.user_id
        data_dic['name'] = poi.name
        data_dic['category'] = poi.category
        data_dic['country'] = poi.country
        data_dic['region'] = poi.region
        data_dic['description'] = poi.description
        data_dic['rating'] = str(get_rating(poi.id))
        data_dic['icon'] = '/static/img/marker/map-mark.png'
        data_dic['lat'] = poi.latitude
        data_dic['lng'] = poi.longitude
        
        map_data.append(data_dic)

    return render_template("home.html", _anchor="where_togo_area", is_authenticated=is_authenticated, user_pois_json=user_pois_dict, pois=map_data, map_init=map_init, search_defaults=search_defaults, categories=get_category_list(), regions=get_region_list(), countries=get_country_list()) 
    # return render_template("home.html", pois=data)


@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/contact")
def contact():
    return render_template("contact.html")

@app.route("/login", methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and bcrypt.check_password_hash(user.password, form.password.data):
            login_user(user, remember=form.remember.data)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('home'))
        else:
            flash('Login Unsuccessful. Please check email and password', 'danger')
    return render_template('login.html', title='Login', form=form)

@app.route("/register", methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    form = RegistrationForm()
    if form.validate_on_submit():
        hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        user = User(username=form.username.data, email=form.email.data, password=hashed_password)
        db.session.add(user)
        db.session.commit()
        flash('Your account has been created! You are now able to log in', 'success')
        return redirect(url_for('login'))

    return render_template("register.html", form=form)

@app.route("/logout")
def logout():
    logout_user()
    return redirect(url_for('home'))

def save_picture(form_picture):
    # randomize the users profile picture so it doesn't conflict with another user
    random_hex = secrets.token_hex(8)
    # f_name, f_ext = os.path.splitext(form_picture.filename) # This returns file name and file extension.  we just want extension
    _, f_ext = os.path.splitext(form_picture.filename) # This returns file name and file extension.  we just want extension
    picture_fn = random_hex + f_ext
    picture_path = os.path.join(app.root_path, 'static/img/profile_pics', picture_fn)

    # form_picture.save(picture_path)
    output_size = (125, 125)
    i = Image.open(form_picture)
    i.thumbnail(output_size)
    i.save(picture_path)

    return picture_fn

@app.route("/account", methods=['GET', 'POST'])
@login_required
def account():
    form = UpdateAccountForm()
    if form.validate_on_submit():
        if form.picture.data:
            picture_file = save_picture(form.picture.data)
            current_user.image_file = picture_file  #image file is what we defined in our models.py
        current_user.username = form.username.data
        current_user.email = form.email.data
        db.session.commit()

        flash('Your account has been updated!', 'success')
        # want to redirect instead of render template so a get request is called instead of  post saying form resubmissions
        return redirect(url_for('account')) 
    elif request.method == 'GET':
        form.username.data = current_user.username
        form.email.data = current_user.email

    image_file = url_for('static', filename='img/profile_pics/' + current_user.image_file)
    return render_template('account.html', title='Account',
                           image_file=image_file, form=form)

@app.route("/poi/new", methods=['GET', 'POST'])
# @login_required
def new_poi():
    form = PoiCreateForm()
    user_id =1    # admin for an anonymous user

    if current_user.is_authenticated:
        user_id=current_user.id

    print("User id is:  ", user_id)

    if form.validate_on_submit():

        # Update POIS table
        poi = Pois(user_id=user_id, name=form.name.data,latitude=float(form.latitude.data), longitude=float(form.longitude.data),
                 region=get_country_region(form.country.data), country=form.country.data, category=form.category.data, description=form.description.data,
                 nearest_icao_code=form.nearest_airport.data)
        
        db.session.add(poi)
        db.session.commit()

        #Update Rating table
        print('Poi ID is: ', poi.id) # This gets the above poi that was just committed.
        rating = Ratings(user_id=user_id, poi_id= poi.id, rating_score=4)
        db.session.add(rating)
        db.session.commit()

        flash('A new point of interest has been created!', 'success')
        return redirect(url_for('home'))
    return render_template('create_poi.html', form=form, legend='New Poi')


@app.route("/poi/<int:poi_id>")
@login_required
def poi(poi_id):
    poi = Pois.query.get_or_404(poi_id)
    return render_template('poi.html', poi=poi)


@app.route("/poi/<int:poi_id>/update", methods=['GET', 'POST'])
@login_required
def update_post(poi_id):
    poi = Pois.query.get_or_404(poi_id)
    print('current poi ID is ', poi.id)
    print(current_user.username)
    if (current_user.username != 'admin') and (poi.user_id != current_user.id):  
        abort(403)

    form = PoiUpdateForm()
    if form.validate_on_submit():
        poi.user_id= current_user.id
        poi.name=form.name.data
        poi.latitude=float(form.latitude.data)
        poi.longitude=float(form.longitude.data)
        poi.region=get_country_region(form.country.data)
        poi.country=form.country.data
        poi.category=form.category.data
        poi.description=form.description.data
        poi.nearest_icao_code=form.nearest_airport.data
        poi.rating=5
        db.session.commit()
        flash('Your post has been updated!', 'success')
        return redirect(url_for('poi', poi_id=poi.id))
    elif request.method == 'GET':
        form.name.data = poi.name
        form.latitude.data = poi.latitude
        form.longitude.data = poi.longitude
        form.country.data = poi.country
        form.description.data = poi.description
        form.nearest_airport.data = poi.nearest_icao_code
        # form.rating.data = poi.rating

    return render_template('update_poi.html', form=form)


@app.route("/poi/<int:poi_id>/delete", methods=['POST'])
@login_required
def delete_post(poi_id):
    print('delete post poid is ', poi_id )
    poi = Pois.query.get_or_404(poi_id)
    if (current_user.username != 'admin') and (poi.user_id != current_user.id):  
        abort(403)
    db.session.delete(poi)
    db.session.commit()
    flash('Your post has been deleted!', 'success')
    return redirect(url_for('home'))

@app.route("/build_db")
@login_required
def build_db():

    # open spreadsheet
    workbook = load_workbook(filename="poi_unedited_ouput.xlsx")
    sheet = workbook.active
    print("######################")
    print(sheet.cell(row=10, column=3).value)

    print('Building dadtabase')

    if (current_user.username == 'admin') and (False):
        
        # Test Create
        user_id = 1 # admin will create all these 

        for count, row in enumerate(sheet.rows, start=1):
            print(count)
            if count ==1:
                continue    # dont include header

            if row[0].value == "":
                break   # no more data in spreadhseet

            

            poi = Pois(
                user_id=user_id,
                name=row[0].value,latitude=float(row[2].value),
                longitude=float(row[3].value),
                region=get_country_region(row[4].value),
                country=row[4].value, category=row[1].value,
                description=row[6].value
            )
        
            db.session.add(poi)
            db.session.commit()

            #Update Rating table
            # print('Poi ID is: ', poi.id) # This gets the above poi that was just committed.
            rating = Ratings(user_id=user_id, poi_id= poi.id, rating_score=4)
            db.session.add(rating)
            db.session.commit()

        flash('Database has been built', 'success')
        return redirect(url_for('home'))
    else:
        
        abort(403)


def get_rating(poi_id):

    sum_rating = 0

    ratings = Ratings.query.filter_by(poi_id=poi_id).all()
    number_of_ratings = 0

    for row in ratings:
        sum_rating += int(row.rating_score)
        number_of_ratings += 1

    return '{0:3.1f}'.format(sum_rating/number_of_ratings)

def get_user_rating(poi_id):

    rating = Ratings.query.filter_by(poi_id=poi_id).first()

    if rating:
        return str(rating.rating_score)

    return ""

def get_favorited(poi_id):

    favorite = Favorites.query.filter_by(poi_id=poi_id).first()

    if favorite:
        return 'True'
    else:
        return 'False'

def get_visited(poi_id):

    visited = Visited.query.filter_by(poi_id=poi_id).first()

    if visited:
        return 'True'
    else:
        return 'False'



